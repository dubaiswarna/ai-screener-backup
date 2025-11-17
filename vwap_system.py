"""
VWAP Flexible Trading System - Core Logic
==========================================
Reusable class for VWAP ladder strategy backtesting
"""

import pandas as pd
import numpy as np
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import warnings
warnings.filterwarnings('ignore')


class VWAPFlexibleSystem:
    """
    Flexible VWAP Ladder Strategy
    Supports optional VWAP, SMA, and Heikin Ashi entries
    """
    
    def __init__(self, max_investment=None, fixed_qty=None, target_percentage=10, threshold_lakhs=5,
                 initial_capital=100000, vwap_enabled=True, sma_period=None, 
                 supertrend_enabled=False, supertrend_period=10, supertrend_multiplier=3, ha_enabled=False,
                 trailing_enabled=False, trailing_percent=5.0, trailing_activation=10.0):
        """
        Initialize with user-defined parameters
        
        Parameters:
        - max_investment: Max investment per day (None if using fixed_qty)
        - fixed_qty: Fixed quantity per order (None if using max_investment)
        - target_percentage: Profit target in % (e.g., 3, 6, 10)
        - threshold_lakhs: Threshold in lakhs (e.g., 3, 4, 5)
        - initial_capital: Starting capital for backtest
        - vwap_enabled: Enable VWAP entries (E3, E4)
        - sma_period: SMA period (enables E5 and E6 entries if set)
        - supertrend_enabled: Enable Supertrend filter
        - ha_enabled: Enable Heikin Ashi (enables E7 and E8 entries if set)
        - trailing_enabled: Enable trailing stop loss
        - trailing_percent: Trailing stop percentage (e.g., 5 = 5% from highest high)
        - trailing_activation: Activate trailing after this profit % reached (e.g., 10 = 10% profit)
        """
        # Validate parameters
        if max_investment is None and fixed_qty is None:
            max_investment = 15000
        
        if fixed_qty is not None:
            fixed_qty = int(fixed_qty)
        
        self.max_investment = max_investment
        self.fixed_qty = fixed_qty
        self.target_percentage = target_percentage
        self.target_price = target_percentage / 100
        self.threshold_lakhs = threshold_lakhs
        self.threshold_amount = threshold_lakhs * 100000
        
        self.initial_capital = initial_capital
        self.position_size = None
        self.profit_target = self.target_price
        
        # Entry configuration
        self.vwap_enabled = vwap_enabled
        self.vwap_entries_enabled = bool(vwap_enabled)
        
        self.sma_period = int(sma_period) if sma_period else None
        self.sma_entries_enabled = bool(self.sma_period)
        
        self.supertrend_enabled = supertrend_enabled
        self.supertrend_period = int(supertrend_period) if supertrend_period else 10
        self.supertrend_multiplier = float(supertrend_multiplier) if supertrend_multiplier else 3
        
        self.ha_enabled = ha_enabled
        self.ha_entries_enabled = bool(ha_enabled)
        
        # Trailing stop loss parameters
        self.trailing_enabled = trailing_enabled
        self.trailing_percent = float(trailing_percent) / 100  # Convert to decimal
        self.trailing_activation = float(trailing_activation) / 100  # Convert to decimal
        
        # Trading parameters
        self.r_low_discount = 0.01
        self.r_vwap_discount = 0.01
        
        # Transaction costs
        self.charges = 0.007
        self.brokerage = 0.003
        self.total_charges = self.charges + self.brokerage
        
        # Above threshold protection
        self.reduced_profit_target = 0.01
        
        # Reset tracking
        self.reset_system()
    
    def reset_system(self):
        """Reset all tracking variables"""
        self.capital = self.initial_capital
        self.data = None
        self.daily_transactions = []
        self.position_summary = []
        self.yearly_summary = []
        
        self.total_shares_held = 0
        self.total_cost = 0
        self.average_cost = 0
        self.target_price_value = 0
        self.position_first_entry_date = None  # Track first buy date for holding period
        
        # Trailing stop loss tracking
        self.trailing_activated = False
        self.highest_high_since_target = 0
        self.trail_stop_price = 0
    
    def load_data_from_dataframe(self, df):
        """Load data from pandas DataFrame"""
        try:
            # Make a copy to avoid modifying original
            df = df.copy()
            
            # Normalize column names
            df.columns = [str(c).strip().lower() for c in df.columns]
            
            # Try to find date column (might be 'date', 'time', 'datetime', etc.)
            date_col = None
            for col in ['date', 'time', 'datetime', 'timestamp']:
                if col in df.columns:
                    date_col = col
                    break
            
            if date_col is None:
                # Try first column if it looks like a date
                first_col = df.columns[0]
                try:
                    pd.to_datetime(df[first_col].iloc[0])
                    date_col = first_col
                except:
                    raise ValueError(f"Could not find date column. Available columns: {', '.join(df.columns)}")
            
            # Rename to 'date'
            if date_col != 'date':
                df = df.rename(columns={date_col: 'date'})
            
            # Parse date
            df['date'] = pd.to_datetime(df['date'], dayfirst=True, errors='coerce')
            df = df.dropna(subset=['date'])
            
            if len(df) == 0:
                raise ValueError("No valid dates found in data")
            
            # Check for required columns
            required_cols = ['high', 'low']
            missing_cols = [col for col in required_cols if col not in df.columns]
            if missing_cols:
                raise ValueError(f"Missing required columns: {', '.join(missing_cols)}. Available: {', '.join(df.columns)}")
            
            # Clean numeric columns
            for col in ['open', 'high', 'low', 'close', 'vwap', 'volume']:
                if col in df.columns:
                    # Remove commas if present
                    if df[col].dtype == 'object':
                        df[col] = df[col].astype(str).str.replace(',', '').str.replace(' ', '')
                    df[col] = pd.to_numeric(df[col], errors='coerce')
            
            # Drop rows with invalid high/low
            df = df.dropna(subset=['high', 'low'])
            
            if len(df) == 0:
                raise ValueError("No valid price data found")
            
            # Set index
            df = df.sort_values('date').set_index('date')
            
            # If no VWAP column, calculate it
            if 'vwap' not in df.columns or df['vwap'].isna().all():
                if 'close' in df.columns:
                    df['vwap'] = df['close']
                else:
                    df['vwap'] = (df['high'] + df['low']) / 2
            
            # Calculate reference levels
            df['r_low'] = df['low'] * (1 - self.r_low_discount)
            df['r_vwap'] = df['vwap'] * (1 - self.r_vwap_discount)
            
            self.data = df
            
            # Calculate SMA if requested
            if self.sma_period and self.sma_period > 0:
                price_col = 'close' if 'close' in self.data.columns else ('vwap' if 'vwap' in self.data.columns else 'high')
                sma_col = f"SMA_{self.sma_period}"
                self.data[sma_col] = self.data[price_col].rolling(window=self.sma_period, min_periods=self.sma_period).mean()
                self._sma_col = sma_col
            
            # Calculate Supertrend if requested
            if self.supertrend_enabled:
                self.calculate_supertrend()
            
            # Calculate Heikin Ashi if requested
            if self.ha_enabled:
                self.calculate_heikin_ashi()
            
            return True
            
        except Exception as e:
            print(f"Error loading data: {str(e)}")
            return False
    
    def calculate_supertrend(self):
        """Calculate Supertrend indicator"""
        if 'close' not in self.data.columns:
            close_col = 'vwap' if 'vwap' in self.data.columns else 'high'
            self.data['close'] = self.data[close_col]
        
        self.data['prev_close'] = self.data['close'].shift(1)
        self.data['tr1'] = self.data['high'] - self.data['low']
        self.data['tr2'] = abs(self.data['high'] - self.data['prev_close'])
        self.data['tr3'] = abs(self.data['low'] - self.data['prev_close'])
        self.data['TR'] = self.data[['tr1', 'tr2', 'tr3']].max(axis=1)
        self.data['ATR'] = self.data['TR'].rolling(window=self.supertrend_period).mean()
        self.data['hl_avg'] = (self.data['high'] + self.data['low']) / 2
        self.data['basic_ub'] = self.data['hl_avg'] + (self.supertrend_multiplier * self.data['ATR'])
        self.data['basic_lb'] = self.data['hl_avg'] - (self.supertrend_multiplier * self.data['ATR'])
        
        self.data['final_ub'] = 0.0
        self.data['final_lb'] = 0.0
        self.data['supertrend'] = 0.0
        
        for i in range(self.supertrend_period, len(self.data)):
            if i == self.supertrend_period:
                self.data.iloc[i, self.data.columns.get_loc('final_ub')] = self.data.iloc[i]['basic_ub']
                self.data.iloc[i, self.data.columns.get_loc('final_lb')] = self.data.iloc[i]['basic_lb']
                self.data.iloc[i, self.data.columns.get_loc('supertrend')] = self.data.iloc[i]['final_ub']
            else:
                if self.data.iloc[i]['basic_ub'] < self.data.iloc[i-1]['final_ub'] or self.data.iloc[i-1]['close'] > self.data.iloc[i-1]['final_ub']:
                    self.data.iloc[i, self.data.columns.get_loc('final_ub')] = self.data.iloc[i]['basic_ub']
                else:
                    self.data.iloc[i, self.data.columns.get_loc('final_ub')] = self.data.iloc[i-1]['final_ub']
                
                if self.data.iloc[i]['basic_lb'] > self.data.iloc[i-1]['final_lb'] or self.data.iloc[i-1]['close'] < self.data.iloc[i-1]['final_lb']:
                    self.data.iloc[i, self.data.columns.get_loc('final_lb')] = self.data.iloc[i]['basic_lb']
                else:
                    self.data.iloc[i, self.data.columns.get_loc('final_lb')] = self.data.iloc[i-1]['final_lb']
                
                if self.data.iloc[i-1]['supertrend'] == self.data.iloc[i-1]['final_ub']:
                    if self.data.iloc[i]['close'] <= self.data.iloc[i]['final_ub']:
                        self.data.iloc[i, self.data.columns.get_loc('supertrend')] = self.data.iloc[i]['final_ub']
                    else:
                        self.data.iloc[i, self.data.columns.get_loc('supertrend')] = self.data.iloc[i]['final_lb']
                else:
                    if self.data.iloc[i]['close'] >= self.data.iloc[i]['final_lb']:
                        self.data.iloc[i, self.data.columns.get_loc('supertrend')] = self.data.iloc[i]['final_lb']
                    else:
                        self.data.iloc[i, self.data.columns.get_loc('supertrend')] = self.data.iloc[i]['final_ub']
        
        self.data.drop(['prev_close', 'tr1', 'tr2', 'tr3', 'TR', 'hl_avg', 'basic_ub', 'basic_lb', 'final_ub', 'final_lb'], axis=1, inplace=True)
    
    def calculate_heikin_ashi(self):
        """Calculate Heikin Ashi candles"""
        if 'open' not in self.data.columns:
            self.data['open'] = self.data['close']
        
        self.data['ha_close'] = (self.data['open'] + self.data['high'] + self.data['low'] + self.data['close']) / 4
        self.data['ha_open'] = 0.0
        self.data['ha_high'] = 0.0
        self.data['ha_low'] = 0.0
        
        for i in range(len(self.data)):
            if i == 0:
                self.data.iloc[i, self.data.columns.get_loc('ha_open')] = (self.data.iloc[i]['open'] + self.data.iloc[i]['close']) / 2
            else:
                self.data.iloc[i, self.data.columns.get_loc('ha_open')] = (self.data.iloc[i-1]['ha_open'] + self.data.iloc[i-1]['ha_close']) / 2
            
            self.data.iloc[i, self.data.columns.get_loc('ha_high')] = max(self.data.iloc[i]['high'], self.data.iloc[i]['ha_open'], self.data.iloc[i]['ha_close'])
            self.data.iloc[i, self.data.columns.get_loc('ha_low')] = min(self.data.iloc[i]['low'], self.data.iloc[i]['ha_open'], self.data.iloc[i]['ha_close'])
    
    def run_backtest(self):
        """Run complete backtest"""
        if self.data is None:
            return False
        
        # Reset
        self.capital = self.initial_capital
        self.daily_transactions = []
        self.yearly_summary = []
        self.total_shares_held = 0
        self.total_cost = 0
        self.average_cost = 0
        self.target_price_value = 0
        
        # Process each day
        for date in self.data.index:
            daily_transaction = self.simulate_daily_trading(date)
            if daily_transaction:
                self.daily_transactions.append(daily_transaction)
        
        # Generate yearly summary
        self.generate_yearly_summary()
        
        return True
    
    def simulate_daily_trading(self, date):
        """Simulate daily trading"""
        if date not in self.data.index:
            return None
        
        row = self.data.loc[date]
        high_price = row['high']
        low_price = row['low']
        date_index = self.data.index.get_loc(date)
        
        # Initialize transaction
        daily_transaction = {
            'date': date,
            'high': high_price,
            'low': low_price,
            'vwap': row.get('vwap', 0),
            'total_buy_qty': 0,
            'total_buy_value': 0,
            'avg_buy_price': 0,
            'sell_qty': 0,
            'sell_price': 0,
            'sell_value': 0,
            'profit': 0,
            'return_pct': 0,
            'execution': 'No',
            'total_shares_held': self.total_shares_held,
            'total_cost': self.total_cost,
            'average_cost': self.average_cost,
            'target_price': self.target_price_value,
            
            # Add indicator values
            'sma': row.get(self._sma_col, 0) if hasattr(self, '_sma_col') and self._sma_col in row.index else 0,
            'supertrend': row.get('supertrend', 0) if self.supertrend_enabled else 0,
            'ha_low': row.get('ha_low', 0) if self.ha_enabled else 0,
            
            # Initialize E1-E8 prices and quantities
            'E1_price': 0, 'E1_qty': 0,
            'E2_price': 0, 'E2_qty': 0,
            'E3_price': 0, 'E3_qty': 0,
            'E4_price': 0, 'E4_qty': 0,
            'E5_price': 0, 'E5_qty': 0,
            'E6_price': 0, 'E6_qty': 0,
            'E7_price': 0, 'E7_qty': 0,
            'E8_price': 0, 'E8_qty': 0,
            
            # Holding period tracking
            'entry_date': self.position_first_entry_date,
            'exit_date': None,
            'holding_days': 0,
        }
        
        # Filter checks
        block_new_buys_due_to_sma = False
        price_col = 'close' if 'close' in self.data.columns else ('vwap' if 'vwap' in self.data.columns else 'high')
        if hasattr(self, '_sma_col') and self._sma_col in row.index:
            current_sma = row[self._sma_col]
            current_price = row[price_col]
            if pd.isna(current_sma) or current_price < current_sma:
                block_new_buys_due_to_sma = True
        
        block_new_buys_due_to_supertrend = False
        if self.supertrend_enabled and 'supertrend' in row.index:
            current_supertrend = row['supertrend']
            current_price = row[price_col]
            if pd.isna(current_supertrend) or current_price > current_supertrend:
                block_new_buys_due_to_supertrend = True
        
        block_new_buys_due_to_threshold = self.total_cost >= self.threshold_amount
        
        # Calculate entry levels ALWAYS (regardless of filters)
        filled_orders = []
        if date_index > 0:
            prev_date = self.data.index[date_index - 1]
            prev_row = self.data.loc[prev_date]
            
            # Calculate entry levels
            e1_price = prev_row['low']
            e2_price = prev_row['low'] * 0.99
            
            entry_levels = {'E1': e1_price, 'E2': e2_price}
            
            # VWAP entries
            if self.vwap_entries_enabled and 'vwap' in prev_row.index:
                prev_vwap = prev_row['vwap']
                if pd.notna(prev_vwap) and prev_vwap > 0:
                    entry_levels['E3'] = prev_vwap
                    entry_levels['E4'] = prev_vwap * 0.99
            
            # SMA entries
            if self.sma_entries_enabled and hasattr(self, '_sma_col') and self._sma_col in prev_row.index:
                prev_sma = prev_row[self._sma_col]
                if pd.notna(prev_sma) and prev_sma > 0:
                    entry_levels['E5'] = prev_sma
                    entry_levels['E6'] = prev_sma * 0.99
            
            # HA entries
            if self.ha_entries_enabled and 'ha_low' in prev_row.index:
                prev_ha_low = prev_row['ha_low']
                if pd.notna(prev_ha_low) and prev_ha_low > 0:
                    entry_levels['E7'] = prev_ha_low
                    entry_levels['E8'] = prev_ha_low * 0.99
            
            # Store entry prices in daily_transaction (ALWAYS)
            for level, price in entry_levels.items():
                daily_transaction[f'{level}_price'] = price
            
            # Process buy orders (only if filters pass)
            if not block_new_buys_due_to_sma and not block_new_buys_due_to_supertrend and not block_new_buys_due_to_threshold:
                # Calculate quantity
                if self.max_investment is not None and self.max_investment > 0:
                    total_entry_price = sum(entry_levels.values())
                    qty_per_order = max(1, int(self.max_investment / total_entry_price))
                elif self.fixed_qty is not None and self.fixed_qty > 0:
                    qty_per_order = int(self.fixed_qty)
                else:
                    qty_per_order = 1
                
                # Check fills
                for level, price in entry_levels.items():
                    if price > 0 and low_price <= price:
                        filled_orders.append({'level': level, 'price': price, 'qty': qty_per_order, 'value': price * qty_per_order})
                        # Store individual entry quantities
                        daily_transaction[f'{level}_qty'] = qty_per_order
        
        # Process filled orders
        if filled_orders:
            total_qty = sum(o['qty'] for o in filled_orders)
            total_value = sum(o['value'] for o in filled_orders)
            avg_price = total_value / total_qty
            
            total_charges_amount = total_value * self.total_charges
            total_cost_with_charges = total_value + total_charges_amount
            
            daily_transaction['total_buy_qty'] = total_qty
            daily_transaction['total_buy_value'] = total_value
            daily_transaction['avg_buy_price'] = avg_price
            
            if (self.total_cost + total_cost_with_charges) <= self.threshold_amount:
                # Track first entry date (for holding period calculation)
                if self.total_shares_held == 0:
                    self.position_first_entry_date = date
                    daily_transaction['entry_date'] = date
                
                self.total_shares_held += total_qty
                self.total_cost += total_cost_with_charges
                self.average_cost = self.total_cost / self.total_shares_held if self.total_shares_held > 0 else 0
                
                if self.total_cost <= self.threshold_amount:
                    self.target_price_value = self.average_cost * (1 + self.profit_target)
                else:
                    self.target_price_value = self.average_cost * (1 + self.reduced_profit_target)
                
                daily_transaction['total_shares_held'] = self.total_shares_held
                daily_transaction['total_cost'] = self.total_cost
                daily_transaction['average_cost'] = self.average_cost
                daily_transaction['target_price'] = self.target_price_value
        
        # Check if we can sell (with trailing stop loss support)
        sell_triggered = False
        sell_price = 0
        sell_type = 'Fixed Target'
        
        if self.total_shares_held > 0:
            # Check if target price reached (for first time)
            if high_price >= self.target_price_value:
                
                # Trailing stop loss logic
                if self.trailing_enabled and not self.trailing_activated:
                    # Check if profit % meets trailing activation threshold
                    current_profit_pct = ((high_price - self.average_cost) / self.average_cost)
                    
                    if current_profit_pct >= self.trailing_activation:
                        # Activate trailing
                        self.trailing_activated = True
                        self.highest_high_since_target = high_price
                        self.trail_stop_price = high_price * (1 - self.trailing_percent)
                        sell_type = 'Trailing Active'
                
                # If trailing not enabled, sell at target
                if not self.trailing_enabled:
                    sell_triggered = True
                    sell_price = self.target_price_value
                    sell_type = 'Fixed Target'
            
            # Update trailing if activated
            if self.trailing_activated:
                # Update highest high
                if high_price > self.highest_high_since_target:
                    self.highest_high_since_target = high_price
                    self.trail_stop_price = high_price * (1 - self.trailing_percent)
                
                # Check if trail stop hit (use low_price for conservative exit)
                if low_price <= self.trail_stop_price:
                    sell_triggered = True
                    sell_price = self.trail_stop_price
                    sell_type = 'Trailing Stop'
        
        # Execute sell if triggered
        if sell_triggered:
            sell_value_gross = sell_price * self.total_shares_held
            sell_charges_amount = sell_value_gross * self.total_charges
            sell_value_net = sell_value_gross - sell_charges_amount
            
            # Calculate holding period
            if self.position_first_entry_date is not None:
                holding_days = (date - self.position_first_entry_date).days
                daily_transaction['exit_date'] = date
                daily_transaction['holding_days'] = holding_days
            
            daily_transaction['sell_qty'] = self.total_shares_held
            daily_transaction['sell_price'] = sell_price
            daily_transaction['sell_value'] = sell_value_gross
            daily_transaction['profit'] = sell_value_net - self.total_cost
            daily_transaction['return_pct'] = (daily_transaction['profit'] / self.total_cost * 100) if self.total_cost > 0 else 0
            daily_transaction['execution'] = 'Sell'
            daily_transaction['sell_type'] = sell_type
            
            self.capital += daily_transaction['profit']
            
            # Reset position
            self.total_shares_held = 0
            self.total_cost = 0
            self.average_cost = 0
            self.target_price_value = 0
            self.position_first_entry_date = None  # Reset for next trade
            self.trailing_activated = False
            self.highest_high_since_target = 0
            self.trail_stop_price = 0
        
        return daily_transaction
    
    def generate_yearly_summary(self):
        """Generate yearly performance summary"""
        if not self.daily_transactions:
            return
        
        df = pd.DataFrame(self.daily_transactions)
        df['year'] = df['date'].dt.year
        
        yearly_data = []
        
        for year in sorted(df['year'].unique()):
            year_df = df[df['year'] == year]
            
            total_buy_qty = year_df['total_buy_qty'].sum()
            total_sell_qty = year_df['sell_qty'].sum()
            total_profit = year_df['profit'].sum()
            total_return_pct = (total_profit / self.initial_capital) * 100 if self.initial_capital > 0 else 0
            
            total_buy_value = year_df['total_buy_value'].sum()
            avg_buy_price = total_buy_value / total_buy_qty if total_buy_qty > 0 else 0
            
            total_sell_value = year_df['sell_value'].sum()
            avg_sell_price = total_sell_value / total_sell_qty if total_sell_qty > 0 else 0
            
            yearly_data.append({
                'Year': year,
                'Total_Buy_Qty': total_buy_qty,
                'Avg_Buy_Price': round(avg_buy_price, 2),
                'Total_Sell_Qty': total_sell_qty,
                'Avg_Sell_Price': round(avg_sell_price, 2),
                'Profit_Booked': round(total_profit, 2),
                'Return_Percentage': round(total_return_pct, 2),
                'Trading_Days': len(year_df),
                'Profitable_Days': len(year_df[year_df['profit'] > 0])
            })
        
        self.yearly_summary = yearly_data
    
    def export_to_bytesio(self):
        """Export results to BytesIO for Streamlit download"""
        output = BytesIO()
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        self.create_daily_transactions_sheet(wb)
        self.create_yearly_summary_sheet(wb)
        self.create_performance_summary_sheet(wb)
        self.create_open_positions_sheet(wb)
        
        wb.save(output)
        output.seek(0)
        
        return output
    
    def create_daily_transactions_sheet(self, wb):
        """Create daily transactions sheet"""
        ws = wb.create_sheet("Daily Transactions")
        
        if not self.daily_transactions:
            return
        
        df = pd.DataFrame(self.daily_transactions)
        
        # Build dynamic columns
        columns = ['Date', 'High', 'Low', 'VWAP']
        
        if self.sma_period:
            columns.append('SMA')
        if self.supertrend_enabled:
            columns.append('Supertrend')
        if self.ha_enabled:
            columns.append('HA_Low')
        
        # Entry prices
        price_cols = ['E1_Price (Low)', 'E2_Price (Low-1%)']
        qty_cols = ['E1_Qty', 'E2_Qty']
        
        if self.vwap_entries_enabled:
            price_cols.extend(['E3_Price (VWAP)', 'E4_Price (VWAP-1%)'])
            qty_cols.extend(['E3_Qty', 'E4_Qty'])
        
        if self.sma_entries_enabled:
            price_cols.extend(['E5_Price (SMA)', 'E6_Price (SMA-1%)'])
            qty_cols.extend(['E5_Qty', 'E6_Qty'])
        
        if self.ha_entries_enabled:
            price_cols.extend(['E7_Price (HA Low)', 'E8_Price (HA Low-1%)'])
            qty_cols.extend(['E7_Qty', 'E8_Qty'])
        
        columns.extend(price_cols)
        columns.extend(qty_cols)
        columns.extend([
            'Total_Buy_Qty', 'Avg_Buy_Price', 'Total_Buy_Value',
            'Sell_Qty', 'Sell_Price', 'Sell_Value',
            'Profit', 'Return_%', 'Execution', 'Sell_Type',
            'Entry_Date', 'Exit_Date', 'Holding_Days',
            'Profit_Target', 'Exceeded_Threshold',
            'Total_Shares_Held', 'Total_Cost', 'Average_Cost', 'Target_Price'
        ])
        
        # Headers
        for col, header in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data rows
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            col = 1
            ws.cell(row=row_idx, column=col, value=row['date'].date())
            col += 1
            ws.cell(row=row_idx, column=col, value=round(row['high'], 2))
            col += 1
            ws.cell(row=row_idx, column=col, value=round(row['low'], 2))
            col += 1
            ws.cell(row=row_idx, column=col, value=round(row['vwap'], 2))
            col += 1
            
            if self.sma_period:
                ws.cell(row=row_idx, column=col, value=round(row.get('sma', 0), 2) if pd.notna(row.get('sma')) else 0)
                col += 1
            if self.supertrend_enabled:
                ws.cell(row=row_idx, column=col, value=round(row.get('supertrend', 0), 2) if pd.notna(row.get('supertrend')) else 0)
                col += 1
            if self.ha_enabled:
                ws.cell(row=row_idx, column=col, value=round(row.get('ha_low', 0), 2) if pd.notna(row.get('ha_low')) else 0)
                col += 1
            
            # Entry prices
            for e in ['E1', 'E2']:
                ws.cell(row=row_idx, column=col, value=round(row.get(f'{e}_price', 0), 2))
                col += 1
            
            if self.vwap_entries_enabled:
                for e in ['E3', 'E4']:
                    ws.cell(row=row_idx, column=col, value=round(row.get(f'{e}_price', 0), 2))
                    col += 1
            
            if self.sma_entries_enabled:
                for e in ['E5', 'E6']:
                    ws.cell(row=row_idx, column=col, value=round(row.get(f'{e}_price', 0), 2))
                    col += 1
            
            if self.ha_entries_enabled:
                for e in ['E7', 'E8']:
                    ws.cell(row=row_idx, column=col, value=round(row.get(f'{e}_price', 0), 2))
                    col += 1
            
            # Entry quantities
            for e in ['E1', 'E2']:
                ws.cell(row=row_idx, column=col, value=row.get(f'{e}_qty', 0))
                col += 1
            
            if self.vwap_entries_enabled:
                for e in ['E3', 'E4']:
                    ws.cell(row=row_idx, column=col, value=row.get(f'{e}_qty', 0))
                    col += 1
            
            if self.sma_entries_enabled:
                for e in ['E5', 'E6']:
                    ws.cell(row=row_idx, column=col, value=row.get(f'{e}_qty', 0))
                    col += 1
            
            if self.ha_entries_enabled:
                for e in ['E7', 'E8']:
                    ws.cell(row=row_idx, column=col, value=row.get(f'{e}_qty', 0))
                    col += 1
            
            # Rest
            ws.cell(row=row_idx, column=col, value=row['total_buy_qty'])
            col += 1
            ws.cell(row=row_idx, column=col, value=round(row['avg_buy_price'], 2))
            col += 1
            ws.cell(row=row_idx, column=col, value=round(row['total_buy_value'], 2))
            col += 1
            ws.cell(row=row_idx, column=col, value=row['sell_qty'])
            col += 1
            ws.cell(row=row_idx, column=col, value=round(row['sell_price'], 2))
            col += 1
            ws.cell(row=row_idx, column=col, value=round(row['sell_value'], 2))
            col += 1
            ws.cell(row=row_idx, column=col, value=round(row['profit'], 2))
            col += 1
            ws.cell(row=row_idx, column=col, value=round(row['return_pct'], 2))
            col += 1
            ws.cell(row=row_idx, column=col, value=row['execution'])
            col += 1
            ws.cell(row=row_idx, column=col, value=row.get('sell_type', 'Fixed Target'))
            col += 1
            
            # Holding period columns
            entry_date = row.get('entry_date')
            exit_date = row.get('exit_date')
            ws.cell(row=row_idx, column=col, value=entry_date.date() if pd.notna(entry_date) and entry_date is not None else '')
            col += 1
            ws.cell(row=row_idx, column=col, value=exit_date.date() if pd.notna(exit_date) and exit_date is not None else '')
            col += 1
            ws.cell(row=row_idx, column=col, value=row.get('holding_days', 0))
            col += 1
            
            ws.cell(row=row_idx, column=col, value=row.get('profit_target_used', f'{self.target_percentage}%'))
            col += 1
            ws.cell(row=row_idx, column=col, value=row.get('exceeded_threshold', 'No'))
            col += 1
            ws.cell(row=row_idx, column=col, value=row['total_shares_held'])
            col += 1
            ws.cell(row=row_idx, column=col, value=round(row['total_cost'], 2))
            col += 1
            ws.cell(row=row_idx, column=col, value=round(row['average_cost'], 2))
            col += 1
            ws.cell(row=row_idx, column=col, value=round(row['target_price'], 2))
        
        # Auto-adjust widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def create_yearly_summary_sheet(self, wb):
        """Create yearly summary sheet"""
        ws = wb.create_sheet("Yearly Summary")
        
        if not self.yearly_summary:
            return
        
        headers = ['Year', 'Total Buy Qty', 'Avg Buy Price', 'Total Sell Qty', 
                  'Avg Sell Price', 'Profit Booked', 'Return %', 'Trading Days', 'Profitable Days']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for row_idx, year_data in enumerate(self.yearly_summary, 2):
            ws.cell(row=row_idx, column=1, value=year_data['Year'])
            ws.cell(row=row_idx, column=2, value=year_data['Total_Buy_Qty'])
            ws.cell(row=row_idx, column=3, value=year_data['Avg_Buy_Price'])
            ws.cell(row=row_idx, column=4, value=year_data['Total_Sell_Qty'])
            ws.cell(row=row_idx, column=5, value=year_data['Avg_Sell_Price'])
            ws.cell(row=row_idx, column=6, value=year_data['Profit_Booked'])
            ws.cell(row=row_idx, column=7, value=year_data['Return_Percentage'])
            ws.cell(row=row_idx, column=8, value=year_data['Trading_Days'])
            ws.cell(row=row_idx, column=9, value=year_data['Profitable_Days'])
        
        # Auto-adjust widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def create_performance_summary_sheet(self, wb):
        """Create performance summary sheet"""
        ws = wb.create_sheet("Performance Summary")
        
        if not self.daily_transactions:
            return
        
        df = pd.DataFrame(self.daily_transactions)
        
        total_trades = len(df[df['profit'] > 0])
        total_profit = df['profit'].sum()
        total_return = (total_profit / self.initial_capital) * 100
        final_capital = self.initial_capital + total_profit
        
        # Count fills
        e5_fills = len(df[df.get('E5_qty', pd.Series([0])) > 0]) if 'E5_qty' in df.columns else 0
        e6_fills = len(df[df.get('E6_qty', pd.Series([0])) > 0]) if 'E6_qty' in df.columns else 0
        e7_fills = len(df[df.get('E7_qty', pd.Series([0])) > 0]) if 'E7_qty' in df.columns else 0
        e8_fills = len(df[df.get('E8_qty', pd.Series([0])) > 0]) if 'E8_qty' in df.columns else 0
        
        # Threshold metrics
        trades_normal = df[(df.get('exceeded_threshold', 'No') == 'No') & (df['execution'] == 'Sell')]
        count_normal = len(trades_normal)
        profit_normal = trades_normal['profit'].sum() if count_normal > 0 else 0
        
        trades_above = df[(df.get('exceeded_threshold', 'No') == 'Yes') & (df['execution'] == 'Sell')]
        count_above = len(trades_above)
        profit_above = trades_above['profit'].sum() if count_above > 0 else 0
        
        metrics = [
            ['Metric', 'Value'],
            ['Strategy', f'VWAP Ladder ({self.target_percentage}% / 1% targets)'],
            ['Target Profit', f'{self.target_percentage}% (below Rs {self.threshold_lakhs}L)'],
            ['Threshold', f'Rs {self.threshold_lakhs}L ({self.threshold_amount:,})'],
            ['Reduced Target', f'1% (above Rs {self.threshold_lakhs}L)'],
            ['', ''],
            ['Initial Capital', f'Rs {self.initial_capital:,}'],
            ['Final Capital', f'Rs {final_capital:,.2f}'],
            ['Total Profit', f'Rs {total_profit:,.2f}'],
            ['Total Return %', f'{total_return:.2f}%'],
            ['', ''],
            ['Total Trades', total_trades],
            [f'  - Trades with {self.target_percentage}% target', f'{count_normal} trades, Rs {profit_normal:,.2f} profit'],
            ['  - Trades with 1% target (>threshold)', f'{count_above} trades, Rs {profit_above:,.2f} profit'],
            ['Win Rate %', '100.0%' if total_trades > 0 else '0%'],
            ['Average Profit per Trade', f'Rs {total_profit/total_trades:.2f}' if total_trades > 0 else 'Rs 0'],
            ['', '']
        ]
        
        # Mode info
        if self.max_investment:
            metrics.extend([
                ['Mode', 'Max Investment (Dynamic Quantity)'],
                ['Max Investment', f'Rs {self.max_investment:,}'],
                ['Quantity Calculation', 'Dynamic (based on price)']
            ])
        else:
            total_e = 2
            if self.vwap_entries_enabled:
                total_e += 2
            if self.sma_entries_enabled:
                total_e += 2
            if self.ha_entries_enabled:
                total_e += 2
            metrics.extend([
                ['Mode', 'Fixed Quantity'],
                ['Fixed Quantity per Order', f'{self.fixed_qty} shares'],
                ['Total Quantity (All Orders)', f'{self.fixed_qty * total_e} shares']
            ])
        
        # Entry points
        total_e = 2
        if self.vwap_entries_enabled:
            total_e += 2
        if self.sma_entries_enabled:
            total_e += 2
        if self.ha_entries_enabled:
            total_e += 2
        
        metrics.extend([
            ['', ''],
            ['Entry Points', f'{total_e} ENTRIES'],
            ['  - E1', 'Previous Day Low'],
            ['  - E2', 'Previous Day Low - 1%']
        ])
        
        if self.vwap_entries_enabled:
            metrics.extend([
                ['  - E3', 'Previous Day VWAP'],
                ['  - E4', 'Previous Day VWAP - 1%']
            ])
        
        if self.sma_entries_enabled:
            metrics.extend([
                ['  - E5', f'Previous Day SMA({self.sma_period})'],
                ['  - E6', f'Previous Day SMA({self.sma_period}) - 1%'],
                ['E5 Fills', f'{e5_fills} times'],
                ['E6 Fills', f'{e6_fills} times']
            ])
        
        if self.ha_entries_enabled:
            metrics.extend([
                ['  - E7', 'Previous Day HA Low'],
                ['  - E8', 'Previous Day HA Low - 1%'],
                ['E7 Fills', f'{e7_fills} times'],
                ['E8 Fills', f'{e8_fills} times']
            ])
        
        metrics.extend([
            ['', ''],
            ['Transaction Charges', '1.0% (0.7% charges + 0.3% brokerage)'],
            ['Strategy Period', f"{self.data.index.min().date()} to {self.data.index.max().date()}"],
            ['Total Trading Days', len(self.data)]
        ])
        
        # Write to sheet
        for row_idx, (metric, value) in enumerate(metrics, 1):
            ws.cell(row=row_idx, column=1, value=metric)
            ws.cell(row=row_idx, column=2, value=value)
            
            if row_idx == 1:
                ws.cell(row=row_idx, column=1).font = Font(bold=True)
                ws.cell(row=row_idx, column=2).font = Font(bold=True)
                ws.cell(row=row_idx, column=1).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                ws.cell(row=row_idx, column=2).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Auto-adjust widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def create_open_positions_sheet(self, wb):
        """Create open positions sheet with live P&L"""
        ws = wb.create_sheet("Open Positions")
        
        if not self.daily_transactions:
            ws.cell(row=1, column=1, value="No transactions available")
            return
        
        df = pd.DataFrame(self.daily_transactions)
        
        # Filter open positions (total_shares_held > 0)
        open_pos_df = df[df['total_shares_held'] > 0].copy()
        
        if open_pos_df.empty:
            ws.cell(row=1, column=1, value="No open positions")
            ws.cell(row=2, column=1, value="All positions have been closed")
            return
        
        # Get the last open segment: rows after the most recent flat period (total_shares_held == 0)
        zero_pos = df[df['total_shares_held'] == 0]
        if not zero_pos.empty:
            last_flat_idx = zero_pos.index.max()
            open_pos_df = open_pos_df[open_pos_df.index > last_flat_idx]
        
        if open_pos_df.empty:
            ws.cell(row=1, column=1, value="No open positions")
            ws.cell(row=2, column=1, value="All positions have been closed")
            return
        
        # Get the last row for current position status
        last_row = open_pos_df.iloc[-1]
        
        # Determine buy date: first non-null entry_date in this segment, else first date with position
        if 'entry_date' in open_pos_df and open_pos_df['entry_date'].notna().any():
            buy_date = open_pos_df.loc[open_pos_df['entry_date'].notna(), 'entry_date'].iloc[0]
        else:
            buy_date = open_pos_df['date'].iloc[0]
        
        # Extract position data
        holding_qty = int(last_row.get('total_shares_held', 0))
        avg_buy_price = float(last_row.get('average_cost', 0.0))
        
        # Current market price (CMP): VWAP if available, else mid of high/low, else avg buy
        vwap_val = float(last_row.get('vwap', 0.0) or 0.0)
        high_val = float(last_row.get('high', 0.0) or 0.0)
        low_val = float(last_row.get('low', 0.0) or 0.0)
        if vwap_val > 0:
            cmp_price = vwap_val
        elif high_val > 0 and low_val > 0:
            cmp_price = (high_val + low_val) / 2.0
        else:
            cmp_price = avg_buy_price
        
        sell_target = float(last_row.get('target_price', 0.0) or 0.0)
        current_date = last_row.get('date')
        
        # Calculate live profit metrics
        investment = avg_buy_price * holding_qty
        running_profit = (cmp_price - avg_buy_price) * holding_qty
        profit_pct_from_cmp = ((cmp_price - avg_buy_price) / avg_buy_price * 100) if avg_buy_price > 0 else 0.0
        pct_to_target_from_cmp = ((sell_target - cmp_price) / cmp_price * 100) if cmp_price > 0 and sell_target > 0 else 0.0
        
        # Calculate holding period and format dates
        try:
            if hasattr(buy_date, 'date'):
                buy_date_val = buy_date.date()
            elif pd.notna(buy_date):
                buy_date_val = pd.to_datetime(buy_date).date()
            else:
                buy_date_val = None
        except:
            buy_date_val = None
        
        try:
            if hasattr(current_date, 'date'):
                current_date_val = current_date.date()
            elif pd.notna(current_date):
                current_date_val = pd.to_datetime(current_date).date()
            else:
                current_date_val = None
        except:
            current_date_val = None
        
        try:
            holding_period = (pd.to_datetime(current_date) - pd.to_datetime(buy_date)).days if pd.notna(buy_date) and pd.notna(current_date) else 0
        except:
            holding_period = 0
        
        # Define columns
        headers = [
            'S.No',
            'Buy Date',
            'Buy Avg Price',
            'Buy Qty',
            'CMP',
            'Sell Target',
            '% of Sell Target from CMP',
            '% Profit from CMP',
            'Running Profit',
            'Holding Period (Days)',
            'Investment Value'
        ]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data row
        row_data = [
            1,
            buy_date_val,
            round(avg_buy_price, 2),
            holding_qty,
            round(cmp_price, 2),
            round(sell_target, 2),
            round(pct_to_target_from_cmp, 2),
            round(profit_pct_from_cmp, 2),
            round(running_profit, 2),
            holding_period,
            round(investment, 2)
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=2, column=col, value=value)
            if col == 2:  # Buy Date column
                if value:
                    cell.value = value
                    cell.number_format = 'yyyy-mm-dd'
            elif col == 3 or col == 5 or col == 6 or col == 9 or col == 11:  # Price/amount columns
                cell.number_format = '0.00'
            elif col == 7 or col == 8:  # Percentage columns (already in % form, e.g., 5.25 = 5.25%)
                cell.number_format = '0.00'
            elif col == 4:  # Quantity (integer)
                cell.number_format = '0'
            elif col == 10:  # Holding Period (integer)
                cell.number_format = '0'
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        if isinstance(cell.value, (int, float)):
                            length = len(f"{cell.value:.2f}")
                        else:
                            length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def get_summary(self):
        """Get performance summary"""
        if not self.daily_transactions:
            return {}
        
        df = pd.DataFrame(self.daily_transactions)
        total_trades = len(df[df['profit'] > 0])
        total_profit = df['profit'].sum()
        total_return = (total_profit / self.initial_capital) * 100
        final_capital = self.initial_capital + total_profit
        
        return {
            'total_trades': total_trades,
            'total_profit': total_profit,
            'total_return': total_return,
            'final_capital': final_capital,
            'win_rate': 100.0 if total_trades > 0 else 0,
            'avg_profit_per_trade': total_profit / total_trades if total_trades > 0 else 0
        }


# ============================================================================
# EFFICIENCY SCORE CALCULATION
# ============================================================================

def calculate_efficiency_score(result_dict):
    """
    Calculate efficiency score for a backtest result
    
    Formula: (Profit / Holding Days) × Win Rate Factor
    
    This rewards:
    - High profit
    - Short holding period (capital turnover)
    - High win rate (risk management)
    
    Parameters:
    -----------
    result_dict : dict
        Must contain: profit, avg_holding_days, win_rate
    
    Returns:
    --------
    float : Efficiency score (higher is better)
    """
    profit = result_dict.get('profit', 0)
    holding_days = result_dict.get('avg_holding_days', 0)
    win_rate = result_dict.get('win_rate', 0)
    
    # Avoid division by zero
    if holding_days == 0 or profit <= 0:
        return 0
    
    # Profit per day (capital efficiency)
    profit_per_day = profit / holding_days
    
    # Win rate factor (1.0 to 2.0 range)
    # 0% win rate → 1.0x, 100% win rate → 2.0x
    win_factor = 1.0 + (win_rate / 100.0)
    
    # Efficiency score
    efficiency_score = profit_per_day * win_factor
    
    return efficiency_score


# ============================================================================
# BATCH OPTIMIZER FUNCTIONS
# ============================================================================

def create_batch_optimizer_excel(batch_results, output_path=None):
    """
    Create comprehensive multi-sheet Excel report for batch optimization
    
    Parameters:
    -----------
    batch_results : list of dict
        Each dict contains:
        {
            'stock_name': str,
            'config_name': str,
            'system': VWAPFlexibleSystem (after backtest),
            'profit': float,
            'trades': int,
            'win_rate': float,
            'avg_holding_days': float
        }
    output_path : str, optional
        If provided, saves to file. Otherwise returns BytesIO
    
    Returns:
    --------
    BytesIO if output_path is None, else None
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Sheet 1: Summary (Best Config per Stock)
    create_best_configs_sheet(wb, batch_results)
    
    # Sheet 2: All Results Comparison
    create_all_results_sheet(wb, batch_results)
    
    # Sheet 3+: Individual Stock Details
    stocks = sorted(set(r['stock_name'] for r in batch_results))
    for stock in stocks:
        stock_results = [r for r in batch_results if r['stock_name'] == stock]
        create_stock_detail_sheet(wb, stock, stock_results)
    
    # Sheet N: Analysis Report
    create_analysis_report_sheet(wb, batch_results)
    
    if output_path:
        wb.save(output_path)
        return None
    else:
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output


def create_best_configs_sheet(wb, batch_results):
    """Create summary sheet showing best configuration for each stock"""
    ws = wb.create_sheet("Best Configurations", 0)
    
    # Group by stock and find best config
    stocks_data = {}
    for result in batch_results:
        stock = result['stock_name']
        if stock not in stocks_data:
            stocks_data[stock] = []
        stocks_data[stock].append(result)
    
    # Headers
    headers = ['Stock', 'Best Configuration', 'Entries', 'Total Profit', 'Trades', 
               'Win Rate %', 'Avg Holding Days', 'Return %', 'Efficiency Score', 'Final Capital']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    row_idx = 2
    for stock in sorted(stocks_data.keys()):
        configs = stocks_data[stock]
        # Find best by efficiency score (Profit/Day × Win Rate)
        best = max(configs, key=lambda x: calculate_efficiency_score(x))
        best_efficiency = calculate_efficiency_score(best)
        
        ws.cell(row=row_idx, column=1, value=stock)
        ws.cell(row=row_idx, column=2, value=best['config_name'])
        ws.cell(row=row_idx, column=3, value=best.get('entries', 0))
        ws.cell(row=row_idx, column=4, value=round(best.get('profit', 0), 2))
        ws.cell(row=row_idx, column=5, value=best.get('trades', 0))
        ws.cell(row=row_idx, column=6, value=round(best.get('win_rate', 0), 2))
        ws.cell(row=row_idx, column=7, value=round(best.get('avg_holding_days', 0), 1))
        ws.cell(row=row_idx, column=8, value=round(best.get('return_pct', 0), 2))
        ws.cell(row=row_idx, column=9, value=round(best_efficiency, 2))
        ws.cell(row=row_idx, column=10, value=round(best.get('final_capital', 0), 2))
        
        row_idx += 1
    
    # Auto-adjust widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 40)


def create_all_results_sheet(wb, batch_results):
    """Create sheet with all stock x config results"""
    ws = wb.create_sheet("All Results")
    
    # Headers
    headers = ['Stock', 'Configuration', 'Entries', 'Profit', 'Trades', 
               'Win Rate %', 'Avg Holding Days', 'Return %', 'Final Capital']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Data rows
    row_idx = 2
    for result in sorted(batch_results, key=lambda x: (x['stock_name'], -x.get('profit', 0))):
        ws.cell(row=row_idx, column=1, value=result['stock_name'])
        ws.cell(row=row_idx, column=2, value=result['config_name'])
        ws.cell(row=row_idx, column=3, value=result.get('entries', 0))
        ws.cell(row=row_idx, column=4, value=round(result.get('profit', 0), 2))
        ws.cell(row=row_idx, column=5, value=result.get('trades', 0))
        ws.cell(row=row_idx, column=6, value=round(result.get('win_rate', 0), 2))
        ws.cell(row=row_idx, column=7, value=round(result.get('avg_holding_days', 0), 1))
        ws.cell(row=row_idx, column=8, value=round(result.get('return_pct', 0), 2))
        ws.cell(row=row_idx, column=9, value=round(result.get('final_capital', 0), 2))
        
        row_idx += 1
    
    # Auto-adjust widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 40)


def create_stock_detail_sheet(wb, stock_name, stock_results):
    """Create detailed sheet for a specific stock showing all config results"""
    # Sanitize sheet name (Excel limits)
    safe_name = stock_name[:25]  # Excel sheet name limit is 31 chars
    ws = wb.create_sheet(f"{safe_name}_Details")
    
    # Find best config for this stock (by efficiency score)
    best_config = max(stock_results, key=lambda x: calculate_efficiency_score(x))
    
    # Title
    ws.cell(row=1, column=1, value=f"Stock: {stock_name}")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    
    ws.cell(row=2, column=1, value=f"Best Configuration: {best_config['config_name']}")
    ws.cell(row=2, column=1).font = Font(bold=True, color="00B050")
    
    # Configuration comparison table
    ws.cell(row=4, column=1, value="Configuration Comparison")
    ws.cell(row=4, column=1).font = Font(bold=True, size=12)
    
    headers = ['Configuration', 'Profit', 'Trades', 'Win Rate %', 'Avg Holding Days', 'Return %']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    row_idx = 6
    for result in sorted(stock_results, key=lambda x: -x.get('profit', 0)):
        ws.cell(row=row_idx, column=1, value=result['config_name'])
        ws.cell(row=row_idx, column=2, value=round(result.get('profit', 0), 2))
        ws.cell(row=row_idx, column=3, value=result.get('trades', 0))
        ws.cell(row=row_idx, column=4, value=round(result.get('win_rate', 0), 2))
        ws.cell(row=row_idx, column=5, value=round(result.get('avg_holding_days', 0), 1))
        ws.cell(row=row_idx, column=6, value=round(result.get('return_pct', 0), 2))
        row_idx += 1
    
    # Best config detailed trades
    ws.cell(row=row_idx + 2, column=1, value=f"Detailed Trades - {best_config['config_name']}")
    ws.cell(row=row_idx + 2, column=1).font = Font(bold=True, size=12)
    
    # Get trades from best system
    best_system = best_config.get('system')
    if best_system and hasattr(best_system, 'daily_transactions'):
        df = pd.DataFrame(best_system.daily_transactions)
        trades_df = df[df['execution'] == 'Sell'].copy()
        
        if not trades_df.empty:
            trade_headers = ['Date', 'Buy Price', 'Sell Price', 'Quantity', 'Profit', 'Return %', 'Holding Days']
            trade_row = row_idx + 3
            for col, header in enumerate(trade_headers, 1):
                cell = ws.cell(row=trade_row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            
            trade_row += 1
            for _, trade in trades_df.iterrows():
                ws.cell(row=trade_row, column=1, value=trade['date'].date() if hasattr(trade['date'], 'date') else trade['date'])
                ws.cell(row=trade_row, column=2, value=round(trade.get('average_cost', 0), 2))
                ws.cell(row=trade_row, column=3, value=round(trade.get('sell_price', 0), 2))
                ws.cell(row=trade_row, column=4, value=trade.get('sell_qty', 0))
                ws.cell(row=trade_row, column=5, value=round(trade.get('profit', 0), 2))
                ws.cell(row=trade_row, column=6, value=round(trade.get('return_pct', 0), 2))
                ws.cell(row=trade_row, column=7, value=trade.get('holding_days', 0))
                trade_row += 1
    
    # Auto-adjust widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 40)


def create_analysis_report_sheet(wb, batch_results):
    """Create analysis sheet with statistics and insights"""
    ws = wb.create_sheet("Analysis Report")
    
    # Title
    ws.cell(row=1, column=1, value="Batch Optimization Analysis")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    
    # Overall statistics
    ws.cell(row=3, column=1, value="Overall Statistics")
    ws.cell(row=3, column=1).font = Font(bold=True, size=12)
    
    total_stocks = len(set(r['stock_name'] for r in batch_results))
    total_configs = len(set(r['config_name'] for r in batch_results))
    total_profit = sum(r.get('profit', 0) for r in batch_results)
    avg_profit = total_profit / len(batch_results) if batch_results else 0
    avg_holding = sum(r.get('avg_holding_days', 0) for r in batch_results) / len(batch_results) if batch_results else 0
    
    stats = [
        ['Total Stocks Analyzed', total_stocks],
        ['Total Configurations Tested', total_configs],
        ['Total Backtests Run', len(batch_results)],
        ['Combined Total Profit', round(total_profit, 2)],
        ['Average Profit per Backtest', round(avg_profit, 2)],
        ['Average Holding Period (days)', round(avg_holding, 1)]
    ]
    
    row_idx = 4
    for stat in stats:
        ws.cell(row=row_idx, column=1, value=stat[0])
        ws.cell(row=row_idx, column=1).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=stat[1])
        row_idx += 1
    
    # Configuration popularity (which config wins most often)
    ws.cell(row=row_idx + 2, column=1, value="Most Successful Configurations")
    ws.cell(row=row_idx + 2, column=1).font = Font(bold=True, size=12)
    
    # Group by stock and find winners
    stocks_data = {}
    for result in batch_results:
        stock = result['stock_name']
        if stock not in stocks_data:
            stocks_data[stock] = []
        stocks_data[stock].append(result)
    
    config_wins = {}
    for stock, configs in stocks_data.items():
        best = max(configs, key=lambda x: calculate_efficiency_score(x))
        config_name = best['config_name']
        config_wins[config_name] = config_wins.get(config_name, 0) + 1
    
    row_idx += 3
    ws.cell(row=row_idx, column=1, value="Configuration")
    ws.cell(row=row_idx, column=1).font = Font(bold=True)
    ws.cell(row=row_idx, column=2, value="Times Best")
    ws.cell(row=row_idx, column=2).font = Font(bold=True)
    ws.cell(row=row_idx, column=3, value="% of Stocks")
    ws.cell(row=row_idx, column=3).font = Font(bold=True)
    
    row_idx += 1
    for config, wins in sorted(config_wins.items(), key=lambda x: -x[1]):
        ws.cell(row=row_idx, column=1, value=config)
        ws.cell(row=row_idx, column=2, value=wins)
        ws.cell(row=row_idx, column=3, value=f"{round(wins/total_stocks*100, 1)}%")
        row_idx += 1
    
    # Auto-adjust widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

